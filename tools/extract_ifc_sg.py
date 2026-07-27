"""Turn the Model Content Requirements workbook into shipped data.

Run with the Blender Python that has Bonsai's wheels on it, since openpyxl
comes from there:

    blender -b --python tools/extract_ifc_sg.py -- <workbook.xlsx> <data-dir>

writing `<data-dir>/ifc_sg.json` and `<data-dir>/delivery/<typology>.json`.

The workbook is ~10 MB of spreadsheet across 26 sheets and is not redistributed
with the add-on: it is a published Singapore standard whose licensing is not
ours to assume, and shipping a 10 MB binary to read the requirement rows would
be poor practice regardless. This produces small JSONs of just the
requirements, which is what gets committed.

The workbook carries two datasets, and both are read:

``IFC+SG Database`` -> ``ifc_sg.json``
    The regulatory IFC+SG requirements for CORENET-X submission. One set for
    every project -- the workbook's own "IFC+SG" presentation sheet restates
    it and has been verified identical, so the Database sheet remains the
    single normalised source. Every stage mark in it is "Y"; nothing optional
    exists to lose.

Per-typology sheets -> ``delivery/<typology>.json``
    The Project Delivery requirements (the workbook's filter calls them "for
    DBC Submission"): what a delivery model should carry per building
    typology. These genuinely differ by typology -- different element
    vocabularies, different parameters -- and carry both "Y" (mandatory) and
    "O" (optional) marks, so both are recorded. The visible "Project
    Typology" sheet with its drop-down filters is derived from these hidden
    sheets; extracting the sheets captures everything the drop-downs can
    show (verified: no marked viewer row names a parameter absent from
    them).

Sheets deliberately not extracted: the comparison studies and working notes
("IFC+SG Study", "01_*"), which restate the same requirements; and "MEP
(consultants only)", a stage-less advisory list of technical attributes with
no requirement marks to extract.

What this deliberately does not produce is a mapping from the standard's
element names to IFC classes. The standard does not contain one -- its own
mechanism for recording the class is a required parameter literally named
`IfcExportAs`, i.e. the modeller declares it. So that mapping is a judgement
call, and it lives in the hand-maintained `ifc_sg_classes.json` and
`delivery_classes.json`, which can be corrected without re-running this.
"""

import json
import os
import sys

#: Column positions in "IFC+SG Database". Checked against the header row on
#: every run, so a reordered workbook fails loudly instead of silently
#: producing requirements attached to the wrong element.
COLUMNS = {
    "element": ("ELEMENT", 0),
    "sub_element": ("SUB-ELEMENT", 1),
    "parameter": ("PARAMETER", 2),
    "category": ("CATEGORY", 3),
    "discipline": ("DISCIPLINE", 4),
}

#: The project stages, in order, and the columns they occupy. The workbook's
#: own header for the fifth reads "(including Fabrication)" because the row
#: above carries "Construction &" -- a merged cell that read_only mode does not
#: reassemble, so the readable names are fixed here.
STAGES = [
    ("conceptual", "Conceptual Design", 5),
    ("schematic", "Schematic / Preliminary Design", 6),
    ("detailed", "Detailed Design / Final Design", 7),
    ("tender", "Tender", 8),
    ("construction", "Construction & Fabrication", 9),
    ("as_built", "As-Built", 10),
    ("operation", "O&M / Asset Information", 11),
]

SHEET = "IFC+SG Database"

#: The hidden per-typology Project Delivery sheets, and the keys and labels
#: they ship under. Same column layout as the Database sheet -- checked
#: against each sheet's header on every run, like the Database's is.
TYPOLOGIES = [
    ("Commercial", "commercial", "Commercial"),
    ("Healthcare", "healthcare", "Healthcare"),
    ("Industrial", "industrial", "Industrial"),
    ("PublicResi", "public_residential", "Public Residential"),
    ("PrivateResi", "private_residential", "Private Residential"),
    ("Retail", "retail", "Retail"),
    ("Infra (Road)", "infra_road", "Infrastructure (Road)"),
    ("Infra (Rail)", "infra_rail", "Infrastructure (Rail)"),
]

#: Rows use this where a sub-element needs no parameters of its own.
NOT_APPLICABLE = "N.A"


def bonsai_site_packages():
    """Where Bonsai unpacks its wheels, so openpyxl can be imported."""
    import bpy

    version = ".".join(str(v) for v in bpy.app.version[:2])
    return os.path.join(
        os.path.expandvars("%APPDATA%"),
        "Blender Foundation",
        "Blender",
        version,
        "extensions",
        ".local",
        "lib",
        f"python3.{sys.version_info.minor}",
        "site-packages",
    )


def load_workbook(path):
    try:
        import openpyxl
    except ImportError:
        sys.path.insert(0, bonsai_site_packages())
        import openpyxl

    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def sheet_rows(workbook, sheet, path):
    if sheet not in workbook.sheetnames:
        raise SystemExit(f"{path} has no {sheet!r} sheet; found {workbook.sheetnames}")
    return list(workbook[sheet].iter_rows(values_only=True))


def text(cell):
    return "" if cell is None else str(cell).replace("\n", " ").strip()


def check_header(header):
    """Fail loudly if the workbook's columns have moved."""
    problems = []
    for _key, (expected, index) in COLUMNS.items():
        found = text(header[index]) if index < len(header) else ""
        if found.upper() != expected.upper():
            problems.append(f"column {index}: expected {expected!r}, found {found!r}")
    for _key, _label, index in STAGES:
        if index >= len(header):
            problems.append(f"column {index}: missing, sheet is too narrow")
    if problems:
        raise SystemExit("Workbook layout has changed:\n  " + "\n  ".join(problems))


def extract(rows):
    check_header(rows[0])

    elements = {}
    counted = 0
    for row in rows[1:]:
        element = text(row[COLUMNS["element"][1]])
        if not element:
            continue
        parameter = text(row[COLUMNS["parameter"][1]])
        # "N.A" means the sub-element is required but carries no parameters of
        # its own. Keeping it would produce a requirement to fill in a field
        # that does not exist.
        if not parameter or parameter.upper() in {NOT_APPLICABLE, "N.A."}:
            continue

        stages = [key for key, _label, index in STAGES if text(row[index]).upper() == "Y"]
        if not stages:
            # Required at no stage is not a requirement.
            continue

        entry = elements.setdefault(element, {"parameters": []})
        entry["parameters"].append(
            {
                "name": parameter,
                "sub_element": text(row[COLUMNS["sub_element"][1]]),
                "discipline": text(row[COLUMNS["discipline"][1]]),
                "stages": stages,
            }
        )
        counted += 1

    for entry in elements.values():
        entry["parameters"].sort(key=lambda p: (p["sub_element"], p["name"]))

    return {
        "source": "IFC+SG Model Content Requirements V2.0, 20 Mar 2026",
        "sheet": SHEET,
        "stages": [{"key": k, "label": label} for k, label, _i in STAGES],
        "elements": elements,
    }, counted


def extract_delivery(rows, key, label):
    """One typology's Project Delivery requirements. Same layout, plus "O".

    Unlike the Database sheet, stage cells here carry "Y" (mandatory) and "O"
    (optional), so each parameter records both -- dropping the optional marks
    would misreport the standard, and promoting them to mandatory would
    over-ask. A handful of cells hold leftover pivot values; only the two
    documented marks count as marks.
    """
    check_header(rows[0])

    elements = {}
    counted = 0
    for row in rows[1:]:
        element = text(row[COLUMNS["element"][1]])
        parameter = text(row[COLUMNS["parameter"][1]])
        if not element or not parameter or parameter.upper() in {NOT_APPLICABLE, "N.A."}:
            continue

        stages = [k for k, _label, index in STAGES if text(row[index]).upper() == "Y"]
        optional = [k for k, _label, index in STAGES if text(row[index]).upper() == "O"]
        if not stages and not optional:
            continue

        entry = elements.setdefault(element, {"parameters": []})
        entry["parameters"].append(
            {
                "name": parameter,
                "sub_element": text(row[COLUMNS["sub_element"][1]]),
                "discipline": text(row[COLUMNS["discipline"][1]]),
                "stages": stages,
                "optional_stages": optional,
            }
        )
        counted += 1

    for entry in elements.values():
        entry["parameters"].sort(key=lambda p: (p["sub_element"], p["name"]))

    return {
        "source": "IFC+SG Model Content Requirements V2.0, 20 Mar 2026",
        "dataset": "Project Delivery (DBC Submission)",
        "typology": key,
        "label": label,
        "stages": [{"key": k, "label": label_} for k, label_, _i in STAGES],
        "elements": elements,
    }, counted


def write_json(data, target):
    os.makedirs(os.path.dirname(os.path.abspath(target)), exist_ok=True)
    with open(target, "w", encoding="utf-8") as handle:
        json.dump(data, handle, indent=1, sort_keys=True, ensure_ascii=False)
        handle.write("\n")


def main():
    args = sys.argv[sys.argv.index("--") + 1 :] if "--" in sys.argv else sys.argv[1:]
    if len(args) != 2:
        raise SystemExit(__doc__)
    source, data_dir = args
    if data_dir.endswith(".json"):
        raise SystemExit("The second argument is now the data directory, not a file: " + __doc__)

    workbook = load_workbook(source)

    data, counted = extract(sheet_rows(workbook, SHEET, source))
    target = os.path.join(data_dir, "ifc_sg.json")
    write_json(data, target)

    print("\n===== IFC+SG EXTRACT =====")
    print(f"elements   : {len(data['elements'])}")
    print(f"parameters : {counted}")
    for name in sorted(data["elements"]):
        params = data["elements"][name]["parameters"]
        subs = len({p["sub_element"] for p in params})
        print(f"  {name:<34} {len(params):>4} parameters across {subs} sub-element(s)")
    print(f"wrote      : {target} ({os.path.getsize(target)} bytes)")

    print("\n===== PROJECT DELIVERY EXTRACT =====")
    for sheet, key, label in TYPOLOGIES:
        delivery, counted = extract_delivery(sheet_rows(workbook, sheet, source), key, label)
        target = os.path.join(data_dir, "delivery", f"{key}.json")
        write_json(delivery, target)
        print(
            f"  {label:<26} {len(delivery['elements']):>2} elements, "
            f"{counted:>5} parameters -> {target} ({os.path.getsize(target)} bytes)"
        )

    # A small index, so the add-on can list typologies without touching the
    # (much larger) per-typology files until one is actually chosen.
    index = {
        "source": "IFC+SG Model Content Requirements V2.0, 20 Mar 2026",
        "typologies": [{"key": key, "label": label} for _sheet, key, label in TYPOLOGIES],
    }
    write_json(index, os.path.join(data_dir, "delivery", "index.json"))
    print("===== END =====\n")


main()
